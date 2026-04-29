import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WEB_DIR = Path(__file__).resolve().parents[1]
ROOT = WEB_DIR.parents[1]
WEB_INDEX_DIR = ROOT / "_web_index"
FIELD_DATA_DIR = WEB_INDEX_DIR / "data"
CHART_TEMPLATE_PATH = FIELD_DATA_DIR / "company_chart_templates.json"
OUT = WEB_DIR / "data" / "gi_web_data.js"


def read_json(path, fallback=None):
    if not path.exists():
        return fallback
    return json.loads(path.read_text(encoding="utf-8"))


def rel(path):
    return path.relative_to(ROOT).as_posix()


def strip_md(value):
    text = str(value or "").replace("`", "").strip()
    text = re.sub(r"\[\[([^\]|]+)(?:\|[^\]]+)?\]\]", r"\1", text)
    return text.strip()


def split_list(value):
    if isinstance(value, list):
        return [item for item in value if item]
    if value in (None, ""):
        return []
    return [strip_md(item) for item in re.split(r"[,;；，]", str(value)) if strip_md(item)]


def clean_scalar(value):
    return str(value or "").strip().strip("\"'")


def parse_frontmatter(text):
    match = re.match(r"^---\r?\n([\s\S]*?)\r?\n---\r?\n?", text)
    if not match:
        return {}
    meta = {}
    current = None
    for line in match.group(1).splitlines():
        item = re.match(r"^\s*-\s+(.*)$", line)
        if item and current:
            if not isinstance(meta.get(current), list):
                meta[current] = []
            meta[current].append(clean_scalar(item.group(1)))
            continue
        kv = re.match(r"^([A-Za-z0-9_]+):\s*(.*)$", line)
        if not kv:
            continue
        current = kv.group(1)
        meta[current] = [] if kv.group(2) == "" else clean_scalar(kv.group(2))
    return meta


def markdown_tables(text):
    tables = []
    current = []
    for raw in text.splitlines():
        line = raw.strip()
        if re.match(r"^\|.*\|$", line):
            if not re.match(r"^\|\s*-", line):
                current.append([cell.strip() for cell in line[1:-1].split("|")])
            continue
        if current:
            tables.append(current)
            current = []
    if current:
        tables.append(current)
    return tables


def find_table(text, required_headers):
    for table in markdown_tables(text):
        headers = table[0] if table else []
        if all(header in headers for header in required_headers):
            return table
    return []


def parse_core_judgments(text):
    match = re.search(r"^##\s+.*核心判断.*$", text, re.M)
    if not match:
        return []
    body = re.split(r"\r?\n##\s+", text[match.end():], maxsplit=1)[0]
    out = []
    for line in body.splitlines():
        item = re.match(r"^\s*\d+\.\s+(.+)$", line)
        if item:
            out.append(item.group(1).strip())
    return out


def rows_to_objects(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    out = []
    for row in rows[1:]:
        obj = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            obj[header] = None if value == "" else value
        out.append(obj)
    return out


def latest_field_inventory_file():
    candidates = list(FIELD_DATA_DIR.glob("*field_inventory*.xlsx"))
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda path: (1 if "with_aggregation" in path.name.lower() else 0, path.stat().st_mtime),
        reverse=True,
    )[0]


def normalize_field_meta(row):
    keys = [
        "field_id",
        "display_name",
        "importance",
        "metric_type",
        "business_scope",
        "segment_scope",
        "parent_field_id",
        "aggregation_group",
        "unit",
        "currency",
        "source_doc",
        "extraction_method",
        "extraction_formula",
        "template_file",
    ]
    meta = {key: row.get(key) for key in keys if row.get(key) not in (None, "")}
    meta["comparable_tags"] = split_list(row.get("comparable_tags"))
    return meta


def load_field_inventory():
    empty = {
        "source_path": None,
        "fields_by_company": {},
        "fields_by_id": {},
        "aggregation_index": {},
        "unique_fields": {},
    }
    file = latest_field_inventory_file()
    if not file:
        return empty

    workbook = load_workbook(file, data_only=True, read_only=True)
    template_rows = rows_to_objects(workbook["financial_template_fields"]) if "financial_template_fields" in workbook.sheetnames else []
    unique_rows = rows_to_objects(workbook["financial_unique_fields"]) if "financial_unique_fields" in workbook.sheetnames else []
    aggregation_rows = rows_to_objects(workbook["aggregation_audit"]) if "aggregation_audit" in workbook.sheetnames else []
    inventory = {**empty, "source_path": rel(file)}

    for row in template_rows:
        company_id = row.get("company_id")
        field_id = row.get("field_id")
        if not company_id or not field_id:
            continue
        meta = normalize_field_meta(row)
        inventory["fields_by_company"].setdefault(company_id, {})[field_id] = meta
        inventory["fields_by_id"].setdefault(field_id, {
            "field_id": field_id,
            "display_names": split_list(row.get("display_name")),
            "metric_types": split_list(row.get("metric_type")),
            "business_scopes": split_list(row.get("business_scope")),
            "segment_scopes": split_list(row.get("segment_scope")),
            "parent_field_ids": split_list(row.get("parent_field_id")),
            "aggregation_groups": split_list(row.get("aggregation_group")),
        })

    for row in unique_rows:
        field_id = row.get("field_id")
        if not field_id:
            continue
        inventory["unique_fields"][field_id] = {
            "field_id": field_id,
            "company_count": row.get("company_count"),
            "companies": split_list(row.get("companies")),
            "display_names": split_list(row.get("display_names")),
            "metric_types": split_list(row.get("metric_types")),
            "business_scopes": split_list(row.get("business_scopes")),
            "segment_scopes": split_list(row.get("segment_scopes")),
            "parent_field_ids": split_list(row.get("parent_field_ids")),
            "aggregation_groups": split_list(row.get("aggregation_groups")),
        }

    for row in aggregation_rows:
        company_id = row.get("company_id")
        parent_field_id = row.get("parent_field_id")
        aggregation_group = row.get("aggregation_group")
        if not company_id or not parent_field_id or not aggregation_group:
            continue
        inventory["aggregation_index"].setdefault(company_id, {}).setdefault(parent_field_id, {})[aggregation_group] = {
            "parent_field_id": parent_field_id,
            "aggregation_group": aggregation_group,
            "component_count": int(row.get("component_count") or 0),
            "component_field_ids": split_list(row.get("component_field_ids")),
            "web_behavior": row.get("web_behavior"),
        }

    return inventory


def load_registries():
    company_path = ROOT / "_registry" / "company_registry.json"
    product_path = ROOT / "_registry" / "product_registry.json"
    company_registry = read_json(company_path, {"companies": {}})
    product_registry = read_json(product_path, {"products": {}})
    return {
        "companies": company_registry.get("companies", {}),
        "products": product_registry.get("products", {}),
        "source_paths": {
            "company_registry": rel(company_path) if company_path.exists() else None,
            "product_registry": rel(product_path) if product_path.exists() else None,
        },
    }


def load_chart_templates():
    raw = read_json(CHART_TEMPLATE_PATH, {})
    if not isinstance(raw, dict):
        return {}
    out = {}
    for company_id, items in raw.items():
        if not isinstance(items, list):
            continue
        normalized = []
        for item in items:
            if not isinstance(item, dict) or not item.get("field_id"):
                continue
            normalized.append({
                "field_id": item.get("field_id"),
                "section": item.get("section") or "公司专属指标",
                "title": item.get("title") or item.get("field_id"),
                "chart_type": item.get("chart_type") or "metric_pair",
                "order": item.get("order") or 999,
                "note": item.get("note") or "",
            })
        out[company_id] = sorted(normalized, key=lambda row: (row.get("order") or 999, row.get("field_id") or ""))
    return out


def merge_field(company_id, field_id, field, inventory):
    meta = inventory["fields_by_company"].get(company_id, {}).get(field_id, {})
    out = {"field_id": field_id, **meta, **field}
    for key in [
        "display_name",
        "importance",
        "metric_type",
        "business_scope",
        "segment_scope",
        "parent_field_id",
        "aggregation_group",
        "unit",
        "currency",
        "source_doc",
        "extraction_method",
        "extraction_formula",
    ]:
        if out.get(key) in (None, "") and meta.get(key) is not None:
            out[key] = meta[key]
    raw_tags = split_list(field.get("comparable_tags"))
    out["comparable_tags"] = raw_tags or split_list(meta.get("comparable_tags"))
    out["field_dictionary_source"] = inventory["source_path"] if meta.get("field_id") else None
    return out


def period_rank(period):
    match = re.match(r"^(\d{4})(Q[1-4]|FY)$", str(period or ""))
    if not match:
        return 0
    suffix_rank = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4, "FY": 5}
    return int(match.group(1)) * 10 + suffix_rank.get(match.group(2), 0)


def load_financial_data(inventory, registries):
    base = ROOT / "_financial_data"
    companies = {}
    for file in base.rglob("*.json"):
        if "_templates" in file.parts:
            continue
        raw = read_json(file)
        company_id = raw.get("company_id") or file.parent.name
        registry = registries["companies"].get(company_id, {})
        fields = {
            field_id: merge_field(company_id, field_id, field, inventory)
            for field_id, field in raw.get("fields", {}).items()
        }
        row = {
            **raw,
            "fields": fields,
            "source_path": rel(file),
            "updated_at": datetime.fromtimestamp(file.stat().st_mtime, timezone.utc).isoformat(),
        }
        companies.setdefault(company_id, {
            "id": company_id,
            "company_id": company_id,
            "name_cn": registry.get("name_cn") or company_id.split("_")[0],
            "name_en": registry.get("name_en") or "",
            "ticker": registry.get("ticker") or "",
            "exchange": registry.get("exchange") or "",
            "tier": registry.get("tier"),
            "aliases": registry.get("aliases") or [],
            "periods": [],
        })
        companies[company_id]["periods"].append(row)

    for company in companies.values():
        company["periods"].sort(key=lambda row: period_rank(row.get("calendar_period")))
    return companies


def parse_briefing(file):
    text = file.read_text(encoding="utf-8")
    meta = parse_frontmatter(text)
    if meta.get("type") != "gameintel_briefing":
        return None
    table = find_table(text, ["field_id", "briefing"])
    headers = table[0] if table else []
    field_index = headers.index("field_id") if "field_id" in headers else -1
    briefing_index = headers.index("briefing") if "briefing" in headers else -1
    field_briefings = {}
    if field_index >= 0 and briefing_index >= 0:
        for row in table[1:]:
            field_id = strip_md(row[field_index] if field_index < len(row) else "")
            note = strip_md(row[briefing_index] if briefing_index < len(row) else "")
            if field_id and note:
                field_briefings[field_id] = {"briefing": note, "source_path": rel(file)}
    return {
        "source_path": rel(file),
        "company_id": meta.get("company_id"),
        "calendar_period": meta.get("calendar_period"),
        "fiscal_period": meta.get("fiscal_period"),
        "core_judgments": parse_core_judgments(text),
        "field_briefings": field_briefings,
        "strategy_keywords": meta.get("strategy_keywords") if isinstance(meta.get("strategy_keywords"), list) else [],
        "human_fill_status": meta.get("human_fill_status") or "",
    }


def parse_quarterly_summary(file):
    text = file.read_text(encoding="utf-8")
    meta = parse_frontmatter(text)
    period = meta.get("calendar_period") or file.stem
    return {"source_path": rel(file), "calendar_period": period, "core_judgments": parse_core_judgments(text)}


def parse_event(file):
    text = file.read_text(encoding="utf-8")
    meta = parse_frontmatter(text)
    if meta.get("type") != "gameintel_event":
        return None
    event = {
        "source_path": rel(file),
        "event_id": meta.get("event_id") or file.stem,
        "event_type": meta.get("event_type") or "",
        "event_name": meta.get("event_name") or file.stem,
        "event_date": meta.get("event_date") or "",
        "calendar_period": meta.get("calendar_period") or "",
        "related_periods": meta.get("related_periods") if isinstance(meta.get("related_periods"), list) else [],
        "company_ids": meta.get("company_ids") if isinstance(meta.get("company_ids"), list) else [],
        "topic_tags": meta.get("topic_tags") if isinstance(meta.get("topic_tags"), list) else [],
        "strategy_keywords": meta.get("strategy_keywords") if isinstance(meta.get("strategy_keywords"), list) else [],
        "field_explanations": {},
        "product_signals": [],
    }

    fields = find_table(text, ["field_id", "official_explanation"])
    headers = fields[0] if fields else []
    if fields and "field_id" in headers and "official_explanation" in headers:
        field_index = headers.index("field_id")
        name_index = headers.index("指标") if "指标" in headers else -1
        explanation_index = headers.index("official_explanation")
        source_index = headers.index("source_doc") if "source_doc" in headers else -1
        for row in fields[1:]:
            field_id = strip_md(row[field_index] if field_index < len(row) else "")
            if not field_id:
                continue
            event["field_explanations"][field_id] = {
                "display_name": strip_md(row[name_index] if 0 <= name_index < len(row) else ""),
                "official_explanation": strip_md(row[explanation_index] if explanation_index < len(row) else ""),
                "source_doc": strip_md(row[source_index] if 0 <= source_index < len(row) else ""),
            }

    products = find_table(text, ["product_id", "signal"])
    headers = products[0] if products else []
    if products and "product_id" in headers and "signal" in headers:
        product_index = headers.index("product_id")
        product_name_index = headers.index("product") if "product" in headers else -1
        signal_index = headers.index("signal")
        source_index = headers.index("source_doc") if "source_doc" in headers else -1
        linked_index = headers.index("linked_fields") if "linked_fields" in headers else -1
        for row in products[1:]:
            event["product_signals"].append({
                "product_id": strip_md(row[product_index] if product_index < len(row) else ""),
                "product": strip_md(row[product_name_index] if 0 <= product_name_index < len(row) else ""),
                "signal": strip_md(row[signal_index] if signal_index < len(row) else ""),
                "source_doc": strip_md(row[source_index] if 0 <= source_index < len(row) else ""),
                "linked_fields": strip_md(row[linked_index] if 0 <= linked_index < len(row) else ""),
            })
    return event


def parse_comparable_tags(file):
    if not file.exists():
        return {}
    text = file.read_text(encoding="utf-8")
    table = find_table(text, ["comparable_tags"])
    headers = table[0] if table else []
    if not table or "comparable_tags" not in headers:
        return {}
    tag_index = headers.index("comparable_tags")
    label_index = headers.index("网页比较项") if "网页比较项" in headers else -1
    type_index = headers.index("源字段类型") if "源字段类型" in headers else -1
    output_index = headers.index("输出") if "输出" in headers else -1
    out = {}
    for row in table[1:]:
        tag_id = strip_md(row[tag_index] if tag_index < len(row) else "")
        if not tag_id:
            continue
        out[tag_id] = {
            "label": strip_md(row[label_index] if 0 <= label_index < len(row) else ""),
            "source_type": strip_md(row[type_index] if 0 <= type_index < len(row) else ""),
            "output": strip_md(row[output_index] if 0 <= output_index < len(row) else ""),
        }
    return out


def attach_briefings_and_explanations(companies, briefings, events):
    for company in companies.values():
        for period in company["periods"]:
            key = f"{company['id']}::{period.get('calendar_period')}"
            briefing = briefings.get(key)
            if briefing:
                period["briefing"] = briefing
            related_events = [
                event for event in events
                if company["id"] in event.get("company_ids", [])
                and (
                    period.get("calendar_period") in event.get("related_periods", [])
                    or event.get("calendar_period") == period.get("calendar_period")
                )
            ]
            period["events"] = [event["event_id"] for event in related_events]
            for field_id, field in period.get("fields", {}).items():
                note = (briefing or {}).get("field_briefings", {}).get(field_id)
                if note:
                    field["field_briefing"] = note
                official = None
                for event in related_events:
                    if field_id in event.get("field_explanations", {}):
                        official = event["field_explanations"][field_id]
                        break
                if official:
                    field["official_explanation"] = official.get("official_explanation")
                    field["explanation_source_doc"] = official.get("source_doc")


def main():
    field_inventory = load_field_inventory()
    registries = load_registries()
    chart_templates = load_chart_templates()
    companies = load_financial_data(field_inventory, registries)

    briefings = {}
    briefings_dir = ROOT / "_briefings"
    if briefings_dir.exists():
        for file in briefings_dir.rglob("*.md"):
            if "_quarterly_summaries" in file.parts:
                continue
            item = parse_briefing(file)
            if item and item.get("company_id") and item.get("calendar_period"):
                briefings[f"{item['company_id']}::{item['calendar_period']}"] = item

    quarterly_summaries = {}
    summary_dir = ROOT / "_briefings" / "_quarterly_summaries"
    if summary_dir.exists():
        for file in summary_dir.rglob("*.md"):
            item = parse_quarterly_summary(file)
            quarterly_summaries[item["calendar_period"]] = item

    events = []
    events_dir = ROOT / "_events"
    if events_dir.exists():
        for file in events_dir.rglob("*.md"):
            item = parse_event(file)
            if item:
                events.append(item)
    events.sort(key=lambda event: str(event.get("event_date") or ""), reverse=True)

    attach_briefings_and_explanations(companies, briefings, events)

    def disclosure_date_for(company_id, period):
        for event in events:
            if (
                event.get("event_type") == "earnings_release"
                and company_id in (event.get("company_ids") or [])
                and (
                    event.get("calendar_period") == period
                    or period in (event.get("related_periods") or [])
                )
            ):
                return event.get("event_date") or ""
        return ""

    financial_updates = sorted(
        [
            {
                "company_id": company["id"],
                "company_name": company["name_cn"],
                "calendar_period": row.get("calendar_period"),
                "source_path": row.get("source_path"),
                "updated_at": row.get("updated_at"),
                "disclosure_date": disclosure_date_for(company["id"], row.get("calendar_period")),
            }
            for company in companies.values()
            for row in company["periods"]
        ],
        key=lambda item: (
            1 if item.get("disclosure_date") else 0,
            str(item.get("disclosure_date") or ""),
            period_rank(item.get("calendar_period")),
        ),
        reverse=True,
    )[:12]

    data = {
        "meta": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_root": rel(ROOT),
            "source_policy": "read_only_generated_from_gameintel_v3",
            "field_inventory_source": field_inventory["source_path"],
            "registry_sources": registries["source_paths"],
            "chart_template_source": rel(CHART_TEMPLATE_PATH) if CHART_TEMPLATE_PATH.exists() else None,
        },
        "companies": companies,
        "chart_templates": chart_templates,
        "briefings": briefings,
        "quarterly_summaries": quarterly_summaries,
        "events": events,
        "comparable_tags": parse_comparable_tags(ROOT / "_financial_data" / "comparable_tags.md"),
        "field_catalog": {
            "source_path": field_inventory["source_path"],
            "fields_by_company": field_inventory["fields_by_company"],
            "unique_fields": field_inventory["unique_fields"],
            "aggregation_index": field_inventory["aggregation_index"],
        },
        "registries": {
            "companies": registries["companies"],
            "products": registries["products"],
        },
        "latest_updates": {
            "financial": financial_updates,
            "events": events[:12],
        },
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "// Auto-generated by scripts/build_data.py. Do not edit manually.\n"
        f"window.GI_WEB_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUT.relative_to(Path.cwd())}")


if __name__ == "__main__":
    main()
